# -*- coding: utf8 -*-
import os
import re
import time
import lxml
import shutil
import sys
import aiohttp
import asyncio
from aiohttp_retry import RetryClient, ExponentialRetry
import aiofiles
import requests
from bs4 import BeautifulSoup
from fake_useragent import UserAgent
from openpyxl import load_workbook
from tqdm import tqdm
import datetime
from PIL import Image, ImageFile


class Parser:

    def __init__(self):
        ua = UserAgent()
        self.headers = {'user_agent': ua.random}
        self.token = ''
        self.secret_key = ''
        self.active_token = ''
        self.active_secret_key = ''
        self.base_url = 'https://ralf.ru'
        self.transliteration_dict = {'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e',
                                     'ё': 'yo', 'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k',
                                     'л': 'l', 'м': 'm', 'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r',
                                     'с': 's', 'т': 't', 'у': 'u', 'ф': 'f', 'х': 'kh', 'ц': 'ts',
                                     'ч': 'ch', 'ш': 'sh', 'щ': 'sch', 'ъ': '', 'ы': 'y', 'ь': '',
                                     'э': 'e', 'ю': 'yu', 'я': 'ya'}
        self.article_numbers = []
        self.links_products = {}
        self.article_imgs = {}
        self.article_save_imgs = {}
        self.read_data1_file = ''
        self.read_data2_file = ''

    def open_token_file(self):
        try:
            with open('token.txt', 'r') as file:
                for i, line in enumerate(file):
                    if i == 0:
                        self.token = line.split('=')[1].strip().split(', ')
                    elif i == 1:
                        self.secret_key = line.split('=')[1].strip().split(', ')
        except Exception:
            print('Не удалось прочитать token или secret_key')
            raise IndexError

    def read_file(self):
        try:
            for file in os.listdir():
                if file[:6] == 'data1.':
                    print(f'Получаю артикул товаров из файла {file}')
                    self.read_data1_file = file
                    self.get_article_number_data1()
                if file[:6] == 'data2.':
                    print(f'Получаю артикул товаров из файла {file}')
                    self.read_data2_file = file
                    self.get_article_number_data2()
        except Exception:
            print('Нет файла с именем data.')
            raise IndexError

    def get_article_number_data1(self):
        try:
            wb = load_workbook(filename=self.read_data1_file)
            sheets = wb.sheetnames
            ws = wb[sheets[0]]

            for row in ws.iter_cols(min_col=2, max_col=2, min_row=2):
                for cell in row:
                    if cell.value is None:
                        continue
                    self.article_numbers.append(cell.value.strip())

            self.article_numbers = list(dict.fromkeys(self.article_numbers))
        except Exception as exc:
            print(f'Ошибка {exc} в чтении табличного документа data1.xlsx')
            with open('error.txt', 'a', encoding='utf-8') as file:

                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в чтении табличного документа data1.xlsm, функция - get_article_number()\n')
            raise IndexError

    def get_article_number_data2(self):
        try:
            wb = load_workbook(filename=self.read_data2_file)
            sheets = wb.sheetnames
            ws = wb[sheets[0]]

            for row in ws.iter_cols(min_col=5, max_col=5, min_row=8):
                for cell in row:
                    if cell.value is None:
                        continue
                    self.article_numbers.append(cell.value.strip())

        except Exception as exc:
            print(f'Ошибка {exc} в чтении табличного документа data2.xlsx')
            with open('error.txt', 'a', encoding='utf-8') as file:

                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в чтении табличного документа data2.xlsm, функция - get_article_number()\n')
            raise IndexError

    # замена букв кириллицы на латиницу
    def transliteration_article(self, article):
        new_str = ''
        for i in range(len(article)):
            if type(article[i]) == str:
                for char in article[i]:
                    if char.lower() in self.transliteration_dict:

                        new_str += self.transliteration_dict[char.lower()]  # замена букв на латинские
                    else:
                        new_str += char
        translated_article = new_str  # замена элемента в списке на новую строку
        return translated_article

    async def get_link_img(self, session, article, translated_article):
        try:

            retry_options = ExponentialRetry(attempts=3)
            retry_client = RetryClient(raise_for_status=False, retry_options=retry_options, client_session=session,
                                       start_timeout=0.5)
            async with retry_client.get(
                    url=f'{self.base_url}/catalog/{translated_article}/') as response:
                if response.ok:
                    sys.stdout.write("\r")
                    sys.stdout.write(f'Получаю ссылку на товар {article}')
                    sys.stdout.flush()

                    resp = await response.text()
                    soup = BeautifulSoup(resp, features='lxml')
                    link_image_found = soup.find_all('div', class_='swiper-slide')
                    images = []
                    for link_image in link_image_found:
                        try:
                            images.append(f"{link_image.find('img')['data-zoom']}")
                        except Exception:
                            continue
                    # берётся 2 и 3 фото
                    self.article_imgs[article] = images[1:3]

        except Exception as exc:
            print(f'Ошибка {exc} в получении ссылок на товары')
            with open('error.txt', 'a', encoding='utf-8') as file:
                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в получении ссылок на товары, функция - get_link_product()\n')

    # можно в адрессную строку сразу подставлять артикул, но меняя буквы на латиницу
    async def get_link_img_run_async(self):
        connector = aiohttp.TCPConnector(force_close=True)
        async with aiohttp.ClientSession(headers=self.headers, connector=connector) as session:
            tasks = []
            for article in self.article_numbers:
                translated_article = self.transliteration_article(article)
                task = asyncio.create_task(self.get_link_img(session, article, translated_article))
                tasks.append(task)
                if len(tasks) % 50 == 0:
                    await asyncio.gather(*tasks)
            await asyncio.gather(*tasks)

    async def save_images(self, session, urls, name_img):
        try:
            images = []

            sys.stdout.write("\r")
            sys.stdout.write(f'Сохраняю изображение для {name_img}')
            sys.stdout.flush()

            for a, url in enumerate(urls):
                date_now = datetime.datetime.now()
                async with aiofiles.open(f'./img/{name_img}_{date_now.strftime("%M%S%f")}_{a}.jpg', mode='wb') as f:
                    async with session.get(f'{self.base_url}{url}') as response:
                        images.append(f'./img/{name_img}_{date_now.strftime("%M%S%f")}_{a}.jpg')
                        async for x in response.content.iter_chunked(1024):
                            await f.write(x)

            self.article_imgs[name_img] = images
        except Exception as exc:
            print(f'Ошибка {exc} в сохранении изображений товаров')
            with open('error.txt', 'a', encoding='utf-8') as file:

                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в сохранении изображений товаров, функция - save_images()\n')

    async def save_images_run_async(self):
        if not os.path.isdir('./img/'):
            os.mkdir('./img/')
        async with aiohttp.ClientSession() as session:
            tasks = []
            for link in self.article_imgs:
                task = asyncio.create_task(self.save_images(session, urls=self.article_imgs[link], name_img=link))
                tasks.append(task)
                await asyncio.gather(*tasks)

    def resize_img(self):
        try:
            ImageFile.LOAD_TRUNCATED_IMAGES = True
            fixed_height = 426
            for img_file in tqdm(os.listdir('./img/')):
                if img_file[-4:] == '.jpg':
                    img = Image.open(f'./img/{img_file}')
                    if img.mode in ("RGBA", "P"):
                        img = img.convert("RGB")
                    height_percent = (fixed_height / float(img.size[1]))
                    width_size = int((float(img.size[0]) * float(height_percent)))
                    new_image = img.resize((width_size, fixed_height))
                    new_image.save(f'./img/{img_file}')
        except Exception as exc:
            print(f'Ошибка {exc} в изменении разрешения изображений')
            with open('error.txt', 'a', encoding='utf-8') as file:
                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в изменении разрешения изображений, функция - resize_img()\n')

    def sending_to_fotohosting(self):
        self.active_token = self.token[0]
        self.active_secret_key = self.secret_key[0]
        headers = {
            'Authorization': f'TOKEN {self.active_token}',
        }
        for img_url in self.article_imgs:

            img_short_link = []

            sys.stdout.write("\r")
            sys.stdout.write(f'Загружаю изображение для - {img_url}')
            sys.stdout.flush()

            img_links = self.article_imgs[img_url]

            for img in img_links:

                try:
                    files = {
                        'image': open(img, 'rb'),
                        'secret_key': (None, self.active_secret_key),
                    }
                    response = requests.post('https://api.imageban.ru/v1', headers=headers, files=files)
                    if response.json()['status'] == 200:
                        img_short_link.append(f"[URL=https://imageban.ru][IMG]{response.json()['data']['link']}"
                                              f"[/IMG][/URL]")
                    else:
                        print(f'Не удалось загрузить {img}')
                        continue
                except KeyError:
                    print(f'{img_url} ошибка загрузки изображения - {response.json()["error"]["message"]}\n')
                    with open('error.txt', 'a', encoding='utf-8') as file:
                        file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                                   f'{img} ошибка загрузки изображения, функция - sending_to_fotohosting()\n')
                    if response.json()["error"]["message"] == 'File reception error':
                        continue
                    elif response.json()["error"]["message"] == \
                            'Exceeded the daily limit of uploaded images for your account':
                        print('Переключение на второй аккаунт')

                        self.active_token = self.token[1]
                        self.active_secret_key = self.secret_key[1]

                        files = {
                            'image': open(img, 'rb'),
                            'secret_key': (None, self.active_secret_key),
                        }
                        response = requests.post('https://api.imageban.ru/v1', headers=headers, files=files)
                        if response.json()['status'] == 200:
                            img_short_link.append(f"[URL=https://imageban.ru][IMG]{response.json()['data']['link']}"
                                                  f"[/IMG][/URL]")
                        else:
                            print(f'Не удалось загрузить {img}')
                    continue
                except FileNotFoundError:
                    continue
                self.article_save_imgs[img_url] = img_short_link

    def write_final_file_data1(self):
        try:
            if not os.path.isdir('./final_data/'):
                os.mkdir('./final_data/')
            columns = ['BB', 'BC', 'BD']
            wb = load_workbook(filename=self.read_data1_file)
            ws = wb.active

            ws['BB1'] = 'Ссылки на фотографии'
            date_now = datetime.datetime.now()
            for article in self.article_save_imgs:
                for i, link in enumerate(self.article_save_imgs[article]):
                    for row in ws.iter_cols(min_col=2, max_col=2, min_row=2):
                        for cell in row:
                            if cell.value.strip() in article:
                                ws[f'{columns[i]}{cell.row}'] = link

            file_name = f'./final_data/data1_final_{date_now.strftime("%d-%m-%y_%H-%M")}.xlsx'
            wb.save(filename=file_name)
            print(f'Файл {file_name} сохранён')
        except Exception as exc:
            print(f'Ошибка {exc} в записи итогового файла')
            with open('error.txt', 'a', encoding='utf-8') as file:
                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в записи итогового файла, функция - write_final_file_data1()\n')

    def write_final_file_data2(self):
        try:
            if not os.path.isdir('./final_data/'):
                os.mkdir('./final_data/')
            columns = ['BG', 'BH', 'BI']
            wb = load_workbook(filename=self.read_data2_file)
            ws = wb.active

            ws['BG7'] = 'Ссылки на фотографии'
            date_now = datetime.datetime.now()
            for article in self.article_save_imgs:
                for i, link in enumerate(self.article_save_imgs[article]):
                    for row in ws.iter_cols(min_col=5, max_col=5, min_row=8):
                        for cell in row:
                            if cell.value.strip() in article:
                                ws[f'{columns[i]}{cell.row}'] = link

            file_name = f'./final_data/data2_final_{date_now.strftime("%d-%m-%y_%H-%M")}.xlsx'
            wb.save(filename=file_name)
            print(f'Файл {file_name} сохранён')
        except Exception as exc:
            print(f'Ошибка {exc} в записи итогового файла')
            with open('error.txt', 'a', encoding='utf-8') as file:
                file.write(f'{datetime.datetime.now().strftime("%d-%m-%y %H:%M")} '
                           f'Ошибка {exc} в записи итогового файла, функция - write_final_file_data2()\n')

    def run(self):
        try:
            # asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
            print('Начало работы')
            self.open_token_file()
            self.read_file()
            print('\rАртикулы получил')
            print('---------------------------\n')
            print('Ищу изображения товаров')
            asyncio.run(self.get_link_img_run_async())
            print('\nИзображения получены')
            print('---------------------------\n')
            print('Скачиваю изображения')
            asyncio.run(self.save_images_run_async())
            print('\nСкачивание завершено')
            print('---------------------------\n')
            print('Измененяю размер изображений')
            self.resize_img()
            print('\rРазмеры изменены')
            print('---------------------------\n')
            print('Загружаю изображения на фотохостинг')
            self.sending_to_fotohosting()
            print('\nЗагрузка завершена')
            print('---------------------------\n')
            print('Записываю в итоговый файл data1_final')
            self.write_final_file_data1()
            print('Записываю в итоговый файл data2_final')
            self.write_final_file_data2()
            print('Работа завершена')
            print('Для выхода нажмите Enter')
            input()
            shutil.rmtree('./img/')
            print('---------------------------\n')
        except Exception as exc:
            print(f'Произошла ошибка {exc}')
            print('Для выхода нажмите Enter')
            input()
            print('---------------------------\n')


def main():
    p = Parser()
    p.run()


if __name__ == '__main__':
    main()
